import pandas as pd
import openpyxl

import os
import sys

from docxtpl import DocxTemplate

sys.path.append(os.getcwd())

from colorama import Fore
from processors.config import template_column_names,source_file_path, prepared_file_path, summary_sent_path, sent_dir, decisions_dir, files_dir, tmp_files_dir, decisions_dir


def depo_cost_parser(s):
    if not s:
        return (False, 'не заполнено')
    
    try:
        # s = str(s)
        s = s.replace(' ', '').replace('\n', ';')
        #print(s)
        l = s.split(';')
        l = [depo_cost.split(':') for depo_cost in l if depo_cost != '']

        #print(l)
        l = [(depo_cost[0].split(','), depo_cost[1]) for depo_cost in l]

        res = {}
        for depos, cost in l:
            for depo in depos:
                res[depo] = float(cost)
        return(True, res)

    except Exception as e:
        return (False, str(e))
    

def source_file_checker(source_file_path, prepared_file_path, summary_sent_path):

    wb = openpyxl.load_workbook(source_file_path)
    sheet_names = wb.sheetnames

    hidden_sheets = [sheet for sheet in sheet_names if wb[sheet].sheet_state == 'hidden']


    sheet_color_list = []
    for sheet in wb.sheetnames:
        a = wb[sheet].sheet_properties.tabColor
        if a:
            sheet_color_list.append((sheet, a.__dict__))
            # print(Fore.MAGENTA, sheet, a.__dict__, Fore.RESET)
        else:
            sheet_color_list.append((sheet, None))
    

            



    wb.close()

    all_sheets_dict = pd.read_excel(source_file_path, sheet_name=None, dtype=str)
    correct_sheets_dict = {} # В этот словарь будут помещенны датафреймы листов, в которых нет ошибок

    #print(all_sheets_dict)

    summary_list = []
   

    for sheet, df in all_sheets_dict.items():
        # print(Fore.MAGENTA, sheet, Fore.RESET)
        df = df.fillna('')


        if sheet in hidden_sheets:
            sheet_state = 'скрытый'
        else:
            sheet_state = '-'

        df_columns = [str(column) for column in df.columns]
        df_columns_to_check = df_columns[:len(template_column_names)+1]# Проверяем, чтобы первые колонки таблице на листе соответствовали шаблону
        
        if not df_columns_to_check == template_column_names:
            # matching_first_columns = f'ШАБЛОН:\n"{ '", "'.join(template_column_names) }"'
            # matching_first_columns = f'НА ЛИСТЕ:\n"{ '", "'.join(df_columns_to_check) }"'
            matching_first_columns = f'ШАБЛОН:\n"{ '", "'.join(template_column_names) }"\nНА ЛИСТЕ:\n"{ '", "'.join(df_columns_to_check) }"'
        else:
            matching_first_columns = '-'

        columns_not_in_df = [template_column for template_column in template_column_names if template_column not in df_columns]
        if columns_not_in_df:
            columns_not_in_df_err = str(columns_not_in_df)[1:-1]
        else:
            columns_not_in_df_err = '-'

        if 'Курс из iSales' in df_columns and len(df)!=0:
            currency_rate_cell_value = df['Курс из iSales'].iloc[0]
            # print(currency_rate_cell_value)
            if currency_rate_cell_value:
                try:
                    float(str(currency_rate_cell_value).replace(',', '.'))
                    currency_rate_err = '-'
                except ValueError:
                    currency_rate_err = f'На листе { sheet } "Курс из iSales" не число: "{currency_rate_cell_value}"' 
            else:
                currency_rate_err = "не заполнено"
        elif 'Курс из iSales' in df_columns and len(df)==0:
            currency_rate_err = "на листе только заголовки"              
        else:
            currency_rate_err = "не заполнено"


        if 'Ставка из iSales' in df_columns and len(df)!=0:
            depo_cost_cell_value = df['Ставка из iSales'].iloc[0]

            depo_cost_parser_res = depo_cost_parser(depo_cost_cell_value)
            
            if depo_cost_parser_res[0]:
                depo_cost_err = '-' 
            else:
                depo_cost_err = depo_cost_parser_res[1]

        elif 'Ставка из iSales' in df_columns and len(df)==0:
            depo_cost_err = "на листе только заголовки" 
        else:
            depo_cost_err = "не заполнено"

        if 'Номер решения ЭС' in df_columns and len(df)!=0:
            desicion_num = df['Номер решения ЭС'].iloc[0]
            
            if desicion_num:
                desicion_num_err = "-"
            else:
                desicion_num_err = "не заполнено"
        elif 'Номер решения ЭС' in df_columns and len(df)==0:
            desicion_num_err = "на листе только заголовки"               
        else:
            desicion_num_err = "нет колонки"                


        if 'Файл с решением ЭС' in df_columns and len(df)!=0:
            desicion_file_name = df['Файл с решением ЭС'].iloc[0]
            # print(Fore.BLUE, sheet, desicion_file_name, Fore.RESET)

            if desicion_file_name != desicion_file_name or desicion_file_name == "":
                desicion_file_name_err = "не заполнено"
            elif not os.path.exists(os.path.join(decisions_dir, str(desicion_file_name))):
                desicion_file_name_err = f"нет файла: {desicion_file_name}" 
            else:
                desicion_file_name_err = "-"
        elif 'Файл с решением ЭС' in df_columns and len(df)==0:
            desicion_file_name_err = "на листе только заголовки"                
        else:
            desicion_file_name_err = "нет колонки"                  

        if 'Комментарий' in df_columns and len(df)!=0:
            
            rem = df['Комментарий'].iloc[0]
            
        elif 'Комментарий' in df_columns and len(df)==0:
            rem = "на листе только заголовки" 
        else:
            rem = "нет колонки"



        if 'Номер контейнера' in df_columns and len(df)!=0:
            conts = df['Номер контейнера']
            conts = conts.dropna()
            conts = conts[conts != '']
            conts_qty = len(conts)
            last_cont = df['Номер контейнера'].iloc[conts_qty-1]

        elif 'Номер контейнера' in df_columns and len(df)==0:
            conts_qty = 0
            last_cont = "на листе только заголовки" 




        if 'Депо сдачи в КНР' in df_columns and len(df)!=0:
            depos = df['Депо сдачи в КНР']
            depos = depos.dropna()
            depos = depos[depos != '']
            depos_qty = len(depos)

        elif 'Депо сдачи в КНР' in df_columns and len(df)==0:   
            depos_qty = "на листе только заголовки"         

        
        if 'Дата актирования' in df_columns and len(df)>1:
            acted = df['Дата актирования']
            acted = acted.dropna()
            acted = acted[acted != '']
            acted_qty = len(acted)


        summary_list.append(
            [sheet,
            sheet_state,
            rem,
            matching_first_columns,
            # columns_not_in_df_err,
            currency_rate_err,
            depo_cost_err,
            desicion_num_err,
            desicion_file_name_err,
            conts_qty,
            last_cont,
            depos_qty,
            acted_qty
            ])
        
        # это блок который создаёт словарь с датафреймами, которые могут добавляться в подготовленные к отправке на актирование
        if (sheet_state == '-' and
            rem[:4] != 'stop',                    # комментарий не начинается со слова "stop"
            matching_first_columns == '-' and
            # columns_not_in_df_err == '-' and
            currency_rate_err == '-' and
            depo_cost_err == '-' and
            desicion_num_err == '-' and
            desicion_file_name_err == '-'
            ):
                        correct_sheets_dict[sheet] = all_sheets_dict[sheet]


    summary_df = pd.DataFrame(
        summary_list,
        columns = ['Заказ',
                   'Скрытый',
                   'Комментарий',
                   'Соответствие первых колонок шаблону',
                   # 'Нет колонок из шаблона',
                   'Курс из iSales',
                   'Ставка из iSales',
                   'Номер решения ЭС',
                   'Файл с решением ЭС',
                   'Контейнеров на листе',
                   'Последний контейнер',
                   'Заполненных депо сдачи',
                   'Актировано']
                   )



    if os.path.exists(prepared_file_path):
        prepared_df = pd.read_excel(
            prepared_file_path,
            dtype=str)
        prepared_df = prepared_df.rename(columns={'Депо сдачи в КНР': 'Подготовленно к передаче на актирование'})
        prepared_grouped = prepared_df.groupby('Заказ')['Подготовленно к передаче на актирование'].count()
    else:
        prepared_grouped = pd.DataFrame(columns=['Заказ','Подготовленно к передаче на актирование'])

    
    if os.path.exists(summary_sent_path):
        sent_df = pd.read_excel(
            summary_sent_path,
            dtype=str)
        sent_df = sent_df.rename(columns={'Депо сдачи в КНР': 'Передано на актирование'})
        sent_grouped = sent_df.groupby('Заказ')['Передано на актирование'].count()
    else:
        sent_grouped = pd.DataFrame(columns=['Заказ','Передано на актирование'])


    check_df = pd.merge(summary_df, prepared_grouped, on='Заказ', how='left')
    check_df = pd.merge(check_df, sent_grouped, on='Заказ', how='left')
    check_df = check_df.fillna({'Подготовленно к передаче на актирование': 0, 'Передано на актирование':0})

    return {'check_df': check_df,
            'sheet_color_list': sheet_color_list,
            'correct_sheets_dict': correct_sheets_dict}    


def get_last_sent_number():
    sent_files = list(os.walk(sent_dir))[0][2]

    if not sent_files:
        return 0
    else:
        last_sent_file = max(sent_files)
        last_number = last_sent_file.split('_')[1]
        return last_number

def init_project():
    # print(os.getcwd().split('\\')[-1], '_' in os.getcwd().split('\\')[-1])

    if '_' in os.getcwd().split('\\')[-1]:
        return False

            
    if not os.path.exists(files_dir):
        os.mkdir(files_dir)
        print(Fore.GREEN, f'В папке проекта создана папка "{ files_dir.split('\\')[-1] }"', Fore.RESET)

    if not os.path.exists(tmp_files_dir):
        os.mkdir(tmp_files_dir)
        print(Fore.GREEN, f'В папке проекта в папке "{ files_dir.split('\\')[-1] }" создана папка "{ tmp_files_dir.split('\\')[-1] }"', Fore.RESET)

    if not os.path.exists(sent_dir):
        os.mkdir(sent_dir)
        print(Fore.GREEN, f'В папке проекта в папке "{ files_dir.split('\\')[-1] }" создана папка "{ sent_dir.split('\\')[-1] }"', Fore.RESET)

    if not os.path.exists(decisions_dir):
        os.mkdir(decisions_dir)
        print(Fore.GREEN, f'В папке проекта в папке "{ files_dir.split('\\')[-1] }" создана папка "{ decisions_dir.split('\\')[-1] }"', Fore.RESET)        


    return True        

########################################

def create_doc_file(prepared_df, sent_file_doc_path):
    # summary_prepared_df = pd.read_excel(prepared_file_path)

    grouped = prepared_df.groupby(['Заказ', 'Номер решения ЭС','Файл с решением ЭС']).agg(
        conts_qty = ('Номер контейнера', 'count'),
        total_in_rub=('Ставка доплаты, руб', 'sum'))


    context = {'orders': []}

    for i in grouped.itertuples():
        context['orders'].append({
            'order': i[0][0],
            'desicion_num': i[0][1],
            'desicion_file_name': i[0][2],
            'conts_qty': i[1],
            'total_in_rub': i[2]
        })

    doc = DocxTemplate(os.path.join(files_dir,"Шаблон СЗ.docx"))
    doc.render(context)
    doc.save(sent_file_doc_path)















if __name__ == '__main__':

    s = """Chengdu, Chongqing, Shanghai: 355;
Huangpu, Dalang, Shilong, Ningbo: 455;
Tianjin, Qingdao: 355;
Dalian: 155;
Changsha, Zhengzhou: 255;
Yiwu, Wuhan: 355;
Hefei: 305;
Taicang, Xi`an: 355;
Xiamen, Shenzhen(Yantian): 455"""
    print(source_file_checker(source_file_path, prepared_file_path, summary_sent_path))