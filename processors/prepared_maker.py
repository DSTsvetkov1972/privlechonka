import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

from config import source_file_path, prepared_file_path, summary_prepared_path, summary_sent_path
from fns import sorce_checker, depo_cost_parser
from colorama import Fore


while os.path.exists(os.path.join('files', '~$подготовлено к передаче на актирование.xlsx')):
    os.startfile(prepared_file_path)
    input('Закройте файл "подготовлено к передаче на актирование.xlsx" и нажмите ввод')

while os.path.exists(os.path.join('files', '~$сводка подготовки к передаче на актирование.xlsx')):
    os.startfile(summary_prepared_path)
    input('Закройте файл "сводка подготовки к передаче на актирование.xlsx" и нажмите ввод')

if os.path.exists(summary_sent_path):
    summary_sent_df = pd.read_excel(summary_sent_path)
    summary_sent_df['Заказ'] = summary_sent_df['Заказ'].apply(str)
else:
    summary_sent_df = pd.DataFrame(columns=['Заказ','Номер контейнера','Депо сдачи в КНР'])


    
sorce_checker_res = sorce_checker(source_file_path, prepared_file_path, summary_sent_path)
correct_sheets_dict = sorce_checker_res['correct_sheets_dict']

prepared_list = []
err_list = []

for sheet, df in correct_sheets_dict.items():
    currency_rate = df['Курс из iSales'].iloc[0]
    depo_cost_str = df['Ставка из iSales'].iloc[0]
    depo_cost_dict = depo_cost_parser(depo_cost_str)[1]

    df['Заказ'] = sheet
    df = pd.merge(df, summary_sent_df, on = ['Заказ','Номер контейнера'], how = 'left')
    df = df[df['Депо сдачи в КНР_x']!=df['Депо сдачи в КНР_y']]

    # print(Fore.YELLOW, sheet, currency_rate, depo_cost_str, Fore.RESET)

    for t in df.itertuples():
        container_number = t[1]
        gate_in_date = t[2]
        depo = t[3]

        cost = depo_cost_dict.get(depo)
        
        if cost:
            cost_in_rub = cost*currency_rate
            prepared_list.append([sheet, container_number, gate_in_date, depo, cost, cost_in_rub])
        else:
            if depo:
                err_list.append((sheet, container_number, depo, 'нет цены для депо'))
            else:
                err_list.append((sheet, container_number, depo, 'ещё не сдан'))

prepared_df = pd.DataFrame(
    prepared_list,
    columns=['Заказ', 'Номер контейнера', 'Дата сдачи в депо КНР', 'Депо сдачи в КНР', 'Цена $', 'Ставка доплаты, руб']) 

err_df = pd.DataFrame(
    err_list,
    columns=['Заказ', 'Номер контейнера', 'Депо сдачи в КНР', 'Ошибка'])



prepared_df.to_excel(prepared_file_path, index=False, sheet_name='Sheet1')
wb = load_workbook(prepared_file_path)

ws = wb['Sheet1']

ws.freeze_panes = 'A2'

# Устанавливаем ширину для конкретной колонки
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 20


for col in 'ABCDEFG':
    cell = ws[f'{ col }1']
    cell.font = Font(bold=True)  # Жирный шрифт
    cell.alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание по центру

wb.save(prepared_file_path)

os.startfile(prepared_file_path)




err_df.to_excel(summary_prepared_path, index=False, sheet_name='Sheet1')


wb = load_workbook(summary_prepared_path)

ws = wb['Sheet1']

ws.freeze_panes = 'A2'

# Устанавливаем ширину для конкретной колонки
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 20


for col in 'ABCD':
    cell = ws[f'{ col }1']
    cell.font = Font(bold=True)  # Жирный шрифт
    cell.alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание по центру

wb.save(summary_prepared_path)

os.startfile(summary_prepared_path)




