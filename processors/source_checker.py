import os
import sys 
sys.path.append(os.getcwd())

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import FilterColumn
import sys
from processors.config import tmp_files_dir

sys.path.append(os.getcwd())
from processors.fns import source_file_checker
from colorama import Fore

from processors.config import source_file_path, prepared_file_path, summary_source_path, summary_sent_path

def source_checker():
    while os.path.exists(os.path.join(tmp_files_dir, '~$сводка по исходному файлу.xlsx')):
        os.startfile(summary_source_path)
        print(Fore.RED, 'Закройте файл "сводка по исходному файлу.xlsx" и нажмите ввод', Fore.RESET, end='')
        input()
        
    sorce_checker_res = source_file_checker(source_file_path, prepared_file_path, summary_sent_path)
    summary_df = sorce_checker_res['check_df']   

    summary_df.to_excel(summary_source_path, index=False, sheet_name='Sheet1')


    wb = load_workbook(summary_source_path)

    ws = wb['Sheet1']
    ws.row_dimensions[1].height = 45

    ws.auto_filter.ref = f"A1:{ get_column_letter(ws.max_column) }{ len(summary_df) }"
    col_filter = FilterColumn(colId=1)  # колонка C (индекс 2)
    # col_filter.filters = Filters(filter=["скрытый"])
    ws.auto_filter.filterColumn.append(col_filter)

    ws.freeze_panes = 'A2'

    # Устанавливаем ширину для конкретной колонки


    # ws.column_dimensions['A'].width = 12
    # ws.column_dimensions['B'].width = 16
    # ws.column_dimensions['C'].width = 22
    # ws.column_dimensions['D'].width = 16
    # ws.column_dimensions['E'].width = 20
    # ws.column_dimensions['F'].width = 22
    # ws.column_dimensions['G'].width = 18
    # ws.column_dimensions['H'].width = 18
    # ws.column_dimensions['I'].width = 18
    # ws.column_dimensions['J'].width = 18
    # ws.column_dimensions['K'].width = 18    
    # ws.column_dimensions['L'].width = 18    


    for col in range(1, ws.max_column+1):
        cell = ws.cell(column=col, row=1)
        cell.font = Font(bold=True)  # Жирный шрифт
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 18


    for col in range(1, ws.max_column+1):
        for row in range(2, ws.max_row+1):
            cell = ws.cell(column=col, row=row)    
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание по центру

    # переносим в сводку цвета ярлыков из исходника
    for row, sheet_color_tuple in enumerate(sorce_checker_res['sheet_color_list'], 2):
        sheet_color_dict= sheet_color_tuple[1]
        if not sheet_color_dict:
            continue
       
        
        if sheet_color_dict['type'] == 'rgb':
            if sheet_color_dict['tint'] != 0:
                cell_fill = PatternFill(
                    fgColor=Color(rgb=sheet_color_dict['rgb'], tint=sheet_color_dict['tint']),
                    fill_type="solid"
                )
            else:
                cell_fill = PatternFill(fgColor=sheet_color_dict['rgb'], fill_type="solid")
        
        elif sheet_color_dict['type'] == 'theme':
            if sheet_color_dict['tint'] != 0:
                cell_fill = PatternFill(
                    fgColor=Color(theme=sheet_color_dict['theme'], tint=sheet_color_dict['tint']),
                    fill_type="solid"
                )
            else:
                cell_fill = PatternFill(
                    fgColor=Color(theme=sheet_color_dict['theme']),
                    fill_type="solid"
                )
        
        elif sheet_color_dict['type'] == 'tint_only':
            # Используем базовый цвет по умолчанию (синий)
            cell_fill = PatternFill(
                fgColor=Color(rgb="4472C4", tint=sheet_color_dict['tint']),
                fill_type="solid"
            )

        elif sheet_color_dict['type'] == 'indexed':
            if sheet_color_dict['tint'] != 0:
                cell_fill = PatternFill(
                    fgColor=Color(indexed=sheet_color_dict['indexed'], tint=sheet_color_dict['tint']),
                    fill_type="solid"
                )
            else:
                cell_fill = PatternFill(
                    fgColor=Color(indexed=sheet_color_dict['indexed']),
                    fill_type="solid"
                )

        for col, column_name in enumerate(summary_df.columns, 1):
            ws_cell = ws.cell(column=col, row=row)
            ws_cell.fill = cell_fill


    wb.save(summary_source_path)

    os.startfile(summary_source_path)

    print(Fore.GREEN, 'Сводка по файлу "свод по каждому заказу.xlsx" подготовлена и будет открыта поверх всех окон на рабочем столе!' ,Fore.RESET)

if __name__ == '__main__':
    source_checker()




